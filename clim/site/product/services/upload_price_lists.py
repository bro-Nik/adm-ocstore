import re
import xlrd
import pandas as pd
from thefuzz import fuzz

from openpyxl import load_workbook
from io import BytesIO

from clim.models import Product


SKIP_SHEETS = ['Оглавление', 'Курс валют ЦБ']
MODEL_KEYWORDS = ['модель']
PRICE_KEYWORDS = ['мрц', 'рекомендованная розничная цена']


def upload_prices(price_file):
    try:
        # Парсим файл
        price_items = parse_price_file(price_file)

        if not price_items:
            return False, {'text': 'Не найдено данных о ценах в файле', 'type': 'warning'}

        unique_matched = comparison_products(price_items)
        return unique_matched, {'text': 'Прайс обработан', 'type': 'info'}

    except Exception as e:
        raise e
        return False, {'text': f'Ошибка обработки файла: {str(e)}', 'type': 'error'}


def parse_price_file(price_file):
    """Парсит файл прайса (xls или xlsx) с учётом объединённых ячеек"""

    # Определяем тип файла по расширению
    filename = price_file.filename.lower()
    if filename.endswith('.xlsx'):
        return parse_xlsx(price_file)
    if filename.endswith('.xls'):
        return parse_xls(price_file)
    raise ValueError("Неподдерживаемый формат файла. Используйте .xlsx или .xls")


def parse_xlsx(file_storage):
    """Обрабатывает xlsx файл из FileStorage"""
    wb = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)
    file_storage.seek(0)  # Сбрасываем позицию чтения файла
    matched_products = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        sheet = wb[sheet_name]
        header_row, model_col, price_col, price_type = find_headers_xlsx(sheet)

        if not header_row:
            continue

        for row in range(header_row + 1, sheet.max_row + 1):
            model_cell = sheet.cell(row=row, column=model_col)
            price_cell = sheet.cell(row=row, column=price_col)

            # Пропускаем пустые строки
            if not model_cell.value and not price_cell.value:
                continue

            model = process_model_cells_xlsx(sheet, model_cell, model_col, price_cell)
            if not model or model.lower() in [*MODEL_KEYWORDS, 'nan', 'none']:
                continue

            price = process_price_cell(price_cell)
            if price is None:
                continue

            matched_products.append({
                'product_name': model,
                'new_price': price,
                'price_type': price_type,
                'sheet': sheet_name
            })

    return matched_products


def parse_xls(file_storage):
    """Обрабатывает xls файл из FileStorage"""
    if xlrd is None:
        raise ImportError("Для работы с .xls файлами требуется библиотека xlrd")
    
    file_content = file_storage.read()
    wb = xlrd.open_workbook(file_contents=file_content)
    matched_products = []

    for sheet_name in wb.sheet_names():
        if sheet_name in SKIP_SHEETS:
            continue

        sheet = wb.sheet_by_name(sheet_name)
        header_row, model_col, price_col, price_type = find_headers_xls(sheet)
        print(header_row, model_col, price_col, price_type)

        if header_row is None or model_col is None or price_col is None:
            continue

        for row in range(header_row + 1, sheet.nrows):
            model_cell = sheet.cell(row, model_col)
            price_cell = sheet.cell(row, price_col)
            
            if not model_cell.value and not price_cell.value:
                continue

            model = process_model_cells_xls(sheet, row, model_col)
            if not model or str(model).lower() in [*MODEL_KEYWORDS, 'nan', 'none']:
                continue

            price = process_price_cell(price_cell)
            if price is None:
                continue

            matched_products.append({
                'product_name': str(model).strip(),
                'new_price': price,
                'price_type': price_type,
                'sheet': sheet_name
            })

    return matched_products


def find_headers_xlsx(sheet):
    header_row = None
    model_col = None
    price_col = None
    price_type = None

    for row in range(1, 11):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                cell_value = str(cell.value).lower()

                # Проверяем ключевые слова для модели
                if model_col is None:
                    for keyword in MODEL_KEYWORDS:
                        if keyword in cell_value:
                            header_row = row
                            model_col = col
                            break

                # Проверяем ключевые слова для цены
                if price_col is None:
                    for keyword in PRICE_KEYWORDS:
                        if keyword in cell_value:
                            price_col = col
                            price_type = keyword.upper()
                            break
        if header_row:
            break

    return header_row, model_col, price_col, price_type


def find_headers_xls(sheet):
    header_row = None
    model_col = None
    price_col = None
    price_type = None

    for row in range(min(10, sheet.nrows)):
        for col in range(sheet.ncols):
            cell = sheet.cell(row, col)
            if cell.value:
                cell_value = str(cell.value).lower()

                # Проверяем ключевые слова для модели
                if model_col is None:
                    for keyword in MODEL_KEYWORDS:
                        if keyword in cell_value:
                            header_row = row
                            model_col = col
                            break

                # Проверяем ключевые слова для цены
                if price_col is None:
                    for keyword in PRICE_KEYWORDS:
                        if keyword in cell_value:
                            price_col = col
                            price_type = keyword.upper()
                            break

        if model_col is not None and price_col is not None:
            break

    return header_row, model_col, price_col, price_type


def process_model_cells_xlsx(sheet, cell, col, price_cell):
    model_parts = []

    for merged_range in sheet.merged_cells.ranges:
        if price_cell.coordinate in merged_range:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                part_cell = sheet.cell(row=r, column=col)
                if part_cell.value:
                    model_parts.append(str(part_cell.value).strip())
            break

    if not model_parts and cell.value:
        model_parts.append(str(cell.value).strip())

    return ' '.join(model_parts).strip()


def process_model_cells_xls(sheet, row, col):
    # В xlrd нет прямой информации об объединенных ячейках, используем эвристику
    model_parts = []
    current_value = sheet.cell(row, col).value
    
    # Проверяем предыдущие строки
    r = row - 1
    while r >= 0 and sheet.cell(r, col).value == current_value:
        r -= 1
    
    # Проверяем следующие строки
    r = row + 1
    while r < sheet.nrows and sheet.cell(r, col).value == current_value:
        r += 1
    
    # Если есть повторяющиеся значения, считаем что это объединенная ячейка
    if r != row + 1:
        model_parts.append(str(current_value).strip())
    else:
        model_parts.append(str(current_value).strip())
    
    return ' '.join(model_parts).strip()


def process_price_cell(cell):
    if not cell.value:
        return None

    price_str = str(cell.value).strip()
    if not price_str or price_str.lower() in ['nan', 'none', 'под заказ', 'в пути']:
        return None

    try:
        price_str = re.sub(r'[^\d.,]', '', price_str.replace(',', '.'))
        return float(price_str)
    except (ValueError, TypeError):
        return None


def comparison_products(price_items):
    # Сопоставляем с товарами в базе
    matched_products = []
    products = Product.query.all()

    for item in price_items:
        # ищем по названию
        if item['product_name']:
            # Удаляем мусор из названия
            clean_name = re.sub(r'[^\w\s]', '', item['product_name'].split('/')[0].split('\\')[0].split('(')[0])
            clean_name = ' '.join(clean_name.split()[:10])  # Берем первые 10 слов

            max_ratio = 0
            product_ratio = None
            for product in products:

                ratio = fuzz.ratio(product.name.lower(), clean_name.lower())

                if max_ratio < ratio > 40:
                    max_ratio = ratio

                    product_ratio = {
                        'product_id': product.product_id,
                        'product_name': product.name,
                        'price_name': clean_name,
                        'current_price': product.price,
                        'current_special_price': min([o.price for o in product.special_offers], default=None),
                        'new_price': item['new_price'],
                        'match_percentage': ratio,
                        'sheet': item['sheet'],
                        'price_type': item['price_type']
                    }

            if product_ratio:
                matched_products.append(product_ratio)

    # Удаляем дубликаты
    seen = set()
    unique_matched = []
    for p in matched_products:
        key = (p['product_id'], p['new_price'], p['price_type'])
        if key not in seen:
            seen.add(key)
            unique_matched.append(p)

    # Сортируем по проценту совпадения
    unique_matched.sort(key=lambda x: x['match_percentage'], reverse=True)

    return unique_matched
